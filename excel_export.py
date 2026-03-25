"""
Экспорт данных в Excel по строгому шаблону
"""
from typing import List
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from models import ParseResult


def create_excel_with_template(results: List[ParseResult], output_file: str):
    """
    Создание Excel файла по строгому шаблону

    Args:
        results: Список ParseResult
        output_file: Путь к выходному файлу
    """
    # Создаем новую книгу
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Кадастровые данные"

    # Определяем порядок колонок строго по шаблону
    columns = [
        "Кадастровый номер объекта недвижимости",  # A
        # Сведения о земельном участке (B-H)
        "Вид объекта недвижимости",  # B
        "Вид земельного участка",  # C
        "Адрес",  # D
        "Площадь декларированная",  # E
        "Вид разрешенного использования",  # F
        "Форма собственности",  # G
        "Кадастровая стоимость",  # H
        # Сведения об объектах (I-U)
        "Вид объекта недвижимости_obj",  # I
        "Кадастровый номер_obj",  # J
        "Назначение",  # K
        "Площадь общая",  # L
        "Форма собственности_obj",  # M
        "Кадастровая стоимость_obj",  # N
        "Удельный показатель кадастровой стоимости",  # O
        "Количество этажей (в том числе подземных)",  # P
        "Количество подземных этажей",  # Q
        "Материал стен",  # R
        "Завершение строительства",  # S
        "Ввод в эксплуатацию",  # T
        "Сведения о культурном наследии",  # U
    ]

    # СТРОКА 1: Группировка (объединенные ячейки)
    ws.row_dimensions[1].height = 25

    # B-H: "Сведения о земельном участке"
    ws.merge_cells('B1:H1')
    cell_b1 = ws['B1']
    cell_b1.value = "Сведения о земельном участке"
    cell_b1.font = Font(bold=True, size=12)
    cell_b1.alignment = Alignment(horizontal='center', vertical='center')
    cell_b1.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

    # I-U: "Сведения об объектах расположенных на земельных участках"
    ws.merge_cells('I1:U1')
    cell_i1 = ws['I1']
    cell_i1.value = "Сведения об объектах расположенных на земельных участках"
    cell_i1.font = Font(bold=True, size=12)
    cell_i1.alignment = Alignment(horizontal='center', vertical='center')
    cell_i1.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

    # СТРОКА 2: Заголовки колонок
    ws.row_dimensions[2].height = 40

    # Заголовки для отображения (без _obj суффиксов)
    headers_display = {
        "Кадастровый номер объекта недвижимости": "Кадастровый номер объекта недвижимости",
        "Вид объекта недвижимости": "Вид объекта недвижимости",
        "Вид земельного участка": "Вид земельного участка",
        "Адрес": "Адрес",
        "Площадь декларированная": "Площадь декларированная",
        "Вид разрешенного использования": "Вид разрешенного использования",
        "Форма собственности": "Форма собственности",
        "Кадастровая стоимость": "Кадастровая стоимость",
        "Вид объекта недвижимости_obj": "Вид объекта недвижимости",
        "Кадастровый номер_obj": "Кадастровый номер",
        "Назначение": "Назначение",
        "Площадь общая": "Площадь общая",
        "Форма собственности_obj": "Форма собственности",
        "Кадастровая стоимость_obj": "Кадастровая стоимость",
        "Удельный показатель кадастровой стоимости": "Удельный показатель кадастровой стоимости",
        "Количество этажей (в том числе подземных)": "Количество этажей (в том числе подземных)",
        "Количество подземных этажей": "Количество подземных этажей",
        "Материал стен": "Материал стен",
        "Завершение строительства": "Завершение строительства",
        "Ввод в эксплуатацию": "Ввод в эксплуатацию",
        "Сведения о культурном наследии": "Сведения о культурном наследии",
    }

    # Записываем заголовки
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = headers_display[col_name]
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Цвет фона в зависимости от секции
        if col_idx == 1:  # A
            cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        elif 2 <= col_idx <= 8:  # B-H (участок)
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        else:  # I-U (объекты)
            cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

    # Добавляем границы
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col_idx in range(1, len(columns) + 1):
        ws.cell(row=1, column=col_idx).border = thin_border
        ws.cell(row=2, column=col_idx).border = thin_border

    # СТРОКИ 3+: Данные
    current_row = 3

    for result in results:
        rows = result.to_excel_rows()

        for row_data in rows:
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = row_data.get(col_name, "-")
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                cell.border = thin_border

            current_row += 1

    # Настройка ширины колонок
    column_widths = {
        'A': 20,  # Кадастровый номер
        'B': 18,  # Вид объекта
        'C': 18,  # Вид участка
        'D': 40,  # Адрес
        'E': 15,  # Площадь
        'F': 35,  # ВРИ
        'G': 20,  # Форма собственности
        'H': 18,  # Стоимость
        'I': 18,  # Вид объекта
        'J': 20,  # Кадастровый номер
        'K': 20,  # Назначение
        'L': 15,  # Площадь
        'M': 20,  # Форма собственности
        'N': 18,  # Стоимость
        'O': 18,  # Удельный показатель
        'P': 15,  # Этажи
        'Q': 12,  # Подземные этажи
        'R': 15,  # Материал стен
        'S': 12,  # Завершение
        'T': 12,  # Ввод
        'U': 35,  # Культурное наследие
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Закрепляем первые 2 строки
    ws.freeze_panes = 'A3'

    # Сохраняем файл
    wb.save(output_file)
    print(f"✅ Excel файл сохранен: {output_file}")
    print(f"   Строк данных: {current_row - 3}")


def test_excel_export():
    """Тест экспорта в Excel"""
    from api_client import NSPDAPIClient
    from models import ParseResult

    print("="*80)
    print("ТЕСТ ЭКСПОРТА В EXCEL ПО ШАБЛОНУ")
    print("="*80)

    client = NSPDAPIClient()

    try:
        # Получаем данные одного участка с объектами
        print("\n📥 Получение данных участка 77:05:0001016:22...")
        result_dict = client.get_full_parcel_info_with_objects("77:05:0001016:22")

        # Ограничиваем до 3 объектов для теста
        objects_data = result_dict['objects_data'][:3]

        parse_result = ParseResult(
            cadastral_number="77:05:0001016:22",
            parcel=result_dict['parcel_data'],
            objects=objects_data,
            status="Успешно"
        )

        print(f"✅ Загружено:")
        print(f"   Участок: {parse_result.parcel.cadastral_number}")
        print(f"   Объектов: {len(parse_result.objects)}")

        # Создаем Excel
        print(f"\n📝 Создание Excel файла...")
        create_excel_with_template([parse_result], "test_export.xlsx")

        print(f"\n✅ Готово! Откройте файл test_export.xlsx")

    except Exception as e:
        print(f"❌ Ошибка: {e}")
        import traceback
        traceback.print_exc()

    finally:
        client.close()


if __name__ == "__main__":
    test_excel_export()
