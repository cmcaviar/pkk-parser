"""
Модуль для работы с Excel файлами
Чтение входных данных и запись результатов
"""
import logging
from pathlib import Path
from typing import List, Optional
from datetime import datetime
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

from config import config
from models import ParseResult

logger = logging.getLogger(__name__)


class ExcelIO:
    """Класс для работы с Excel файлами"""

    def __init__(self, file_path: str):
        """
        Args:
            file_path: Путь к Excel файлу
        """
        self.file_path = Path(file_path)
        self.workbook: Optional[Workbook] = None

        if not self.file_path.exists():
            raise FileNotFoundError(f"Файл не найден: {self.file_path}")

        logger.info(f"Инициализация ExcelIO для файла: {self.file_path}")

    def read_cadastral_numbers(self) -> List[str]:
        """
        Чтение кадастровых номеров из входного листа

        Returns:
            Список кадастровых номеров
        """
        try:
            logger.info(f"Чтение данных из листа '{config.excel.input_sheet}'")

            # Загрузка книги
            wb = load_workbook(self.file_path, read_only=True, data_only=True)

            # Проверка наличия листа
            if config.excel.input_sheet not in wb.sheetnames:
                raise ValueError(f"Лист '{config.excel.input_sheet}' не найден в файле")

            ws = wb[config.excel.input_sheet]

            # Поиск колонки с кадастровыми номерами
            column_index = None
            header_row = 1

            for cell in ws[header_row]:
                if cell.value and config.excel.input_column.lower() in str(cell.value).lower():
                    column_index = cell.column
                    logger.info(f"Колонка '{config.excel.input_column}' найдена: {get_column_letter(column_index)}")
                    break

            if column_index is None:
                raise ValueError(f"Колонка '{config.excel.input_column}' не найдена в листе '{config.excel.input_sheet}'")

            # Чтение кадастровых номеров
            cadastral_numbers = []

            for row in ws.iter_rows(min_row=header_row + 1, min_col=column_index, max_col=column_index):
                cell = row[0]
                if cell.value:
                    value = str(cell.value).strip()
                    if value:  # Пропускаем пустые строки
                        cadastral_numbers.append(value)

            wb.close()

            logger.info(f"Прочитано кадастровых номеров: {len(cadastral_numbers)}")
            return cadastral_numbers

        except Exception as e:
            logger.error(f"Ошибка чтения данных из Excel: {e}")
            raise

    def write_results(
        self,
        results: List[ParseResult],
        output_file: Optional[str] = None
    ):
        """
        Запись результатов парсинга в Excel

        Args:
            results: Список результатов парсинга
            output_file: Путь к выходному файлу (если None - используется исходный файл)
        """
        try:
            # Определение выходного файла
            if output_file:
                out_path = Path(output_file)
            else:
                # Используем исходный файл - добавляем лист "Результат"
                out_path = self.file_path

            logger.info(f"Запись результатов в файл: {out_path}")

            # Загрузка или создание книги
            if out_path.exists() and out_path == self.file_path:
                # Работаем с исходным файлом - добавляем новый лист
                wb = load_workbook(out_path)

                # Удаляем старый лист с результатами если есть
                if config.excel.output_sheet in wb.sheetnames:
                    logger.info(f"Удаление существующего листа '{config.excel.output_sheet}'")
                    wb.remove(wb[config.excel.output_sheet])

                # Создаем новый лист
                ws = wb.create_sheet(config.excel.output_sheet)

            else:
                # Создаем новый файл
                wb = Workbook()
                ws = wb.active
                ws.title = config.excel.output_sheet

            # Создание заголовков
            self._write_headers(ws)

            # Запись данных
            row_num = 2  # Начинаем со второй строки (первая - заголовки)

            for result in results:
                excel_rows = result.to_excel_rows()

                for excel_row in excel_rows:
                    # Записываем данные в порядке колонок
                    col_num = 1

                    for col_name in config.excel.all_columns:
                        value = excel_row.get(col_name, "Нет данных")
                        ws.cell(row=row_num, column=col_num, value=value)
                        col_num += 1

                    row_num += 1

            # Автоширина колонок
            self._auto_size_columns(ws)

            # Сохранение файла
            wb.save(out_path)
            wb.close()

            logger.info(f"Результаты успешно записаны: {out_path}")
            logger.info(f"Записано строк: {row_num - 2}")

        except Exception as e:
            logger.error(f"Ошибка записи результатов в Excel: {e}")
            raise

    def _write_headers(self, worksheet):
        """
        Запись заголовков таблицы

        Args:
            worksheet: Рабочий лист openpyxl
        """
        # Стили для заголовков
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Запись заголовков
        for idx, col_name in enumerate(config.excel.all_columns, start=1):
            cell = worksheet.cell(row=1, column=idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Замораживание первой строки
        worksheet.freeze_panes = "A2"

    def _auto_size_columns(self, worksheet):
        """
        Автоматическая настройка ширины колонок

        Args:
            worksheet: Рабочий лист openpyxl
        """
        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter

            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Максимум 50 символов
            worksheet.column_dimensions[column_letter].width = adjusted_width

    @staticmethod
    def create_output_filename(base_path: Optional[str] = None) -> str:
        """
        Создание имени выходного файла с датой

        Args:
            base_path: Базовый путь (директория исходного файла)

        Returns:
            Полный путь к выходному файлу
        """
        date_str = datetime.now().strftime("%Y-%m-%d")
        filename = config.excel.output_file_pattern.format(date=date_str)

        if base_path:
            return str(Path(base_path) / filename)
        else:
            return filename


def read_cadastral_numbers_from_excel(excel_file: str) -> List[str]:
    """
    Вспомогательная функция для чтения кадастровых номеров

    Args:
        excel_file: Путь к Excel файлу

    Returns:
        Список кадастровых номеров
    """
    excel_io = ExcelIO(excel_file)
    return excel_io.read_cadastral_numbers()


def write_results_to_excel(
    results: List[ParseResult],
    excel_file: str,
    create_new_file: bool = False
):
    """
    Вспомогательная функция для записи результатов

    Args:
        results: Список результатов
        excel_file: Путь к исходному Excel файлу
        create_new_file: Создать новый файл с датой вместо добавления листа
    """
    excel_io = ExcelIO(excel_file)

    if create_new_file:
        # Создаем новый файл в той же директории
        output_file = ExcelIO.create_output_filename(Path(excel_file).parent)
        excel_io.write_results(results, output_file)
    else:
        # Добавляем лист в исходный файл
        excel_io.write_results(results)
